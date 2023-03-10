import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

CORN_FLOWER_BLUE = "6495ED"
POWDER_BLUE = "B0E0E6"

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


class Report:

    def __init__(self, path):
        self.path = path
        self.csv_filename = path / "data/WebDevelopment.csv"
        self.filename_out = path / "data/WebDevelopment.xlsx"

        self.df = self.__create_df_from_file(self.csv_filename)
        self.wb = Workbook()

    def __get_data_from_csv_file(self, filename):
        return pd.read_csv(filename)

    def __create_df_from_file(self, filename):
        return self.__get_data_from_csv_file(filename)

    def create_first_worksheet(self):
        ws = self.wb.active
        ws.title = "AllWebCourses"
        rows = dataframe_to_rows(self.df, index=False)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    def create_tech_worksheet(self, tech, tech_row_name):
        new_df = self.df[self.df[tech_row_name].str.contains(tech)]
        ws = self.wb.create_sheet(title=tech)
        rows = dataframe_to_rows(new_df, index=False)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                current_cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1:
                    self.format_body_tech_cell(current_cell)

    def format_body_tech_cell(self, cell):
        cell.fill = PatternFill(start_color=POWDER_BLUE,
                                end_color=POWDER_BLUE,
                                fill_type='solid')
        cell.border = THIN_BORDER

    def start(self):
        self.create_first_worksheet()

    def finish(self):
        self.wb.save(self.filename_out)
